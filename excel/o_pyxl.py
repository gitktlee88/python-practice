import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
#openpyxl.__version__

# import os
# os.chdir('C:/Users/Tiger KT/jupyter-notebook')

def as_text(value):
    if value is None:
        return ""
    return str(value)

wb  = openpyxl.load_workbook('Sample_data.xlsx')
# type(wb)

# wb.get_sheet_names()
wb.sheetnames

# sheet = wb.get_sheet_by_name('Sheet1')
ws = wb['Sheet1']
ws['B1'].font = Font(sz=14, bold=True, italic=True)
wb.save('Sample_data')

def adjust_column_width_from_col(ws, min_row, min_col, max_col):

        column_widths = []

        for i, col in enumerate(
                    ws.iter_cols(min_col=min_col, max_col=max_col, min_row=min_row)
                ):

            for cell in col:
                value = cell.value
                if value is not None:
                    print(value)

                    if isinstance(value, str) is False:
                        value = str(value)

                    try:
                        column_widths[i] = max(column_widths[i], len(value))
                    except IndexError:
                        column_widths.append(len(value))

#         print(column_widths)

        for i, width in enumerate(column_widths):

            col_name = get_column_letter(min_col + i)
            value = column_widths[i] + 2
            ws.column_dimensions[col_name].width = value
            print(col_name, value)


if __name__ == '__main__':

    adjust_column_width_from_col(ws, 1,1, ws.max_column)
